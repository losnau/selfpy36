#-*- coding:utf-8
import  requests
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bs4 import  BeautifulSoup
soup=BeautifulSoup(open('test.txt'),'html.parser')
items=soup.find_all('tr')
data={
    'site_id':'10',
    'contentarea_by':'Linux',
    'sort_by':'',
    'sort_order':'',
    'start':'0',
    'end':'99',
    'topic_by':'',
    'product_by':'',
    'type_by':u'所有类别',
    'show_abstract':'true',
    'search_by':'',
    'industry_by':'',
    'series_title_by':'',
}
others='site_id=10&contentarea_by=Linux&sort_by=&sort_order=2&topic_by=&product_by=&type_by=%E6%89%80%E6%9C%89%E7%B1%BB%E5%88%AB&show_abstract=false&search_by=&industry_by=&series_title_by='
base_url='https://www.ibm.com/developerworks/cn/views/linux/libraryview.jsp'
print(type(items)) 
#保存excel
# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()
# 获取当前活跃的worksheet,默认就是第一个worksheet
ws = wb.active
ws.title = u'运维派经验链接'
# 文本对齐方式
align = Alignment(horizontal='center', vertical='center')
# 字体大小
font = Font(size=10)
# 设置表格列宽
ws.column_dimensions["A"].width = 40.0
ws.column_dimensions["B"].width = 50.0
ws.column_dimensions["C"].width = 30.0
ws.column_dimensions["D"].width = 10.0
row = 1  # 行的初始值
column = 1  # 列的初始值

for i in range(0,18):
    # st_en='start='+str(i*100)+'&end='+str((i+1)*100-1)
    # url=base_url+st_en+others
    # print url
    print('----------------------------第'+row+'行') 
    data['start']=str(i*100)
    data['end'] =str((i+1)*100-1)
    response=requests.get(base_url,data)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.find_all('tr')
    for item in items[1:]:
        column = 1  # 列的初始值
        #标题，每执行一次列加1
        article_title=item.td.strong.string
        ws.cell(row=row, column=column).value = article_title
        column=column+1
        #简介      and符号，讲解参考https://my.oschina.net/chuangspace/blog/410833
        article_abstract=item.div.string and item.div.string.strip()
        ws.cell(row=row, column=column).value = article_abstract
        column=column+1
        #日期
        article_href=item.td.a['href']
        ws.cell(row=row, column=column).value = article_href
        column=column+1
        #文章日期
        article_date=item.td.next_sibling.next_sibling.next_sibling.next_sibling.string
        ws.cell(row=row, column=column).value = article_date
        column=column+1
        # 行+1
        row = row + 1

wb.save('ibm.xlsx')
