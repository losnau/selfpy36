#-*- coding:utf-8
from bs4 import  BeautifulSoup
import bs4
import requests
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class JINGYAN:
    def __init__(self):
        self.base_url="http://www.yunweipai.com/category/jingyan"
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0",
            "Host": "www.yunweipai.com",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3",
            "Accept-Encoding": "gzip, deflate",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Cache-Control": "max-age=0",
        }

    def get_content(self, url):
        try:
            response = requests.get(url, headers=self.headers)
            return response.content
        except requests.HTTPError as e:
            print('错误：' + e)
    def get_page(self):
        content=self.get_content(self.base_url)
        soup = BeautifulSoup(content, 'html.parser', from_encoding='utf-8')
        div = soup.find_all('div', class_='pagination')
        return int(div[0].contents[-1]['href'].split('/')[-1])
    def save_excel(self,name):
        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        # 获取当前活跃的worksheet,默认就是第一个worksheet
        ws = wb.active
        ws.title = u'运维派经验链接'
        # 文本对齐方式
        align = Alignment(horizontal='center', vertical='center')
        # 字体大小
        font = Font(size=10)
        # 设置表格列宽
        ws.column_dimensions["A"].width = 40.0
        ws.column_dimensions["B"].width = 50.0

        row = 1         #行的初始值
        column = 1      #列的初始值
        page=self.get_page()
        for i in range(1,page+1):
            print("保存------------第%s页------------" %i )
            ws.cell(row=row, column=column).alignment = align
            ws.cell(row=row, column=column).font = font
            ws.cell(row=row, column=column).value = "------------第%s页------------"%i
            row=row+1
            url=self.base_url+'/page/'+str(i)
            #print('url:'+url)             #测试
            #continue
            content=self.get_content(url)
            soup = BeautifulSoup(content, 'html.parser')
            data = soup.find_all('div', class_='title')
            for title in data:                                  #每循环一次增加一行
                # print(title.a['href']+'-------',end='')               #第一列
                print(u'--第'+str(i)+u'页--'+u'--'+str(row)+str(column)+':'+title.a['href'])
                ws.cell(row=row, column=1).alignment = align
                ws.cell(row=row, column=1).font = font
                ws.cell(row=row, column=1).value = title.a['href']
                if type(title.a.string) == bs4.element.Comment:
                    print(title.a.string)
                else:
                    for title_name in title.stripped_strings:
                        #print(title_name.strip())                       #第二列
                        print('--第'+str(i)+ u'页--'+ u'--'+ str(row)+str(column)+':'+title_name.strip())
                        ws.cell(row=row, column=2).alignment = align
                        ws.cell(row=row, column=2).font = font
                        ws.cell(row=row, column=2).value = title_name.strip()
                        row = row + 1
        wb.save(name)
jy=JINGYAN()
name='yunweipai.xlsx'
jy.save_excel(name)
